"""
Saldenliste Parser Service
Parses Saldenliste CSV/Excel files into structured account data.
Supports multiple formats: BMD, RZL, and custom formats.
"""
import csv
import io
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from enum import Enum
from pathlib import Path
from typing import List, Optional, Dict, Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class SaldenlisteFormat(str, Enum):
    """Supported Saldenliste file formats."""

    BMD = "bmd"
    RZL = "rzl"
    CUSTOM = "custom"
    UNKNOWN = "unknown"


@dataclass
class AccountBalance:
    """Individual account balance entry."""

    account_number: str
    account_name: str
    debit_balance: Optional[Decimal] = None
    credit_balance: Optional[Decimal] = None
    balance: Decimal = Decimal("0")
    kontenklasse: Optional[int] = None  # Mapped from account number


@dataclass
class SaldenlisteData:
    """Structured data from a Saldenliste file."""

    tax_year: int
    company_name: Optional[str] = None
    accounts: List[AccountBalance] = field(default_factory=list)
    total_assets: Optional[Decimal] = None
    total_liabilities: Optional[Decimal] = None
    confidence: float = 0.0
    format_detected: SaldenlisteFormat = SaldenlisteFormat.UNKNOWN


class SaldenlisteParser:
    """Parse Saldenliste CSV/Excel files into structured account data."""

    # Common column name variations for account number
    ACCOUNT_NUMBER_COLUMNS = [
        "konto",
        "kontonummer",
        "account",
        "account_number",
        "accountnumber",
        "kto",
        "kto_nr",
        "kto-nr",
    ]

    # Common column name variations for account name
    ACCOUNT_NAME_COLUMNS = [
        "bezeichnung",
        "kontobezeichnung",
        "name",
        "account_name",
        "accountname",
        "description",
        "konto_bezeichnung",
    ]

    # Common column name variations for balance/amount
    BALANCE_COLUMNS = [
        "saldo",
        "balance",
        "betrag",
        "amount",
        "endsaldo",
        "end_balance",
    ]

    # Debit/Credit columns
    DEBIT_COLUMNS = ["soll", "debit", "sollsaldo", "debit_balance"]
    CREDIT_COLUMNS = ["haben", "credit", "habensaldo", "credit_balance"]
    CSV_ENCODINGS = ("utf-8", "utf-8-sig", "cp1252", "latin-1")

    def parse_csv(self, file_path: str, tax_year: int) -> SaldenlisteData:
        """
        Parse CSV format Saldenliste.

        Args:
            file_path: Path to the CSV file
            tax_year: Tax year for the Saldenliste

        Returns:
            SaldenlisteData with parsed accounts

        Raises:
            ValueError: If file cannot be parsed or required columns are missing
        """
        path = Path(file_path)
        if not path.exists():
            raise ValueError(f"File not found: {file_path}")

        try:
            reader = self._open_csv_reader(file_path)
            if not reader.fieldnames:
                raise ValueError("CSV file has no headers")

            # Normalize column names (lowercase, strip whitespace)
            normalized_headers = {
                col.lower().strip().replace(" ", "_"): col for col in reader.fieldnames
            }

            # Detect format based on headers
            format_detected = self._detect_format_from_headers(list(normalized_headers.keys()))

            # Find column mappings
            account_col = self._find_column(normalized_headers, self.ACCOUNT_NUMBER_COLUMNS)
            name_col = self._find_column(normalized_headers, self.ACCOUNT_NAME_COLUMNS)

            if not account_col or not name_col:
                raise ValueError(
                    f"Required columns not found. Need account number and name columns. "
                    f"Found headers: {list(reader.fieldnames)}"
                )

            # Try to find balance columns (either single balance or debit/credit)
            balance_col = self._find_column(normalized_headers, self.BALANCE_COLUMNS)
            debit_col = self._find_column(normalized_headers, self.DEBIT_COLUMNS)
            credit_col = self._find_column(normalized_headers, self.CREDIT_COLUMNS)

            if not balance_col and not (debit_col or credit_col):
                raise ValueError(
                    "No balance columns found. Need either 'balance' or 'debit'/'credit' columns"
                )

            accounts = []
            for row in reader:
                account = self._parse_csv_row(
                    row, account_col, name_col, balance_col, debit_col, credit_col
                )
                if account:
                    accounts.append(account)

            return SaldenlisteData(
                tax_year=tax_year,
                accounts=accounts,
                confidence=self._calculate_confidence(accounts),
                format_detected=format_detected,
            )

        except Exception as e:
            raise ValueError(f"Failed to parse CSV file: {str(e)}") from e

    def parse_excel(self, file_path: str, tax_year: int) -> SaldenlisteData:
        """
        Parse Excel format Saldenliste using openpyxl.

        Args:
            file_path: Path to the Excel file
            tax_year: Tax year for the Saldenliste

        Returns:
            SaldenlisteData with parsed accounts

        Raises:
            ValueError: If file cannot be parsed or required columns are missing
        """
        path = Path(file_path)
        if not path.exists():
            raise ValueError(f"File not found: {file_path}")

        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet = workbook.active

            if not sheet:
                raise ValueError("Excel file has no active sheet")

            # Find header row (usually first row, but could be offset)
            header_row = self._find_header_row(sheet)
            if not header_row:
                raise ValueError("Could not find header row in Excel file")

            headers = [
                str(cell.value).lower().strip().replace(" ", "_")
                if cell.value
                else f"col_{idx}"
                for idx, cell in enumerate(header_row)
            ]

            # Detect format
            format_detected = self._detect_format_from_headers(headers)

            # Create column mapping
            normalized_headers = {headers[i]: i for i in range(len(headers))}

            # Find column indices
            account_col_idx = self._find_column_index(normalized_headers, self.ACCOUNT_NUMBER_COLUMNS)
            name_col_idx = self._find_column_index(normalized_headers, self.ACCOUNT_NAME_COLUMNS)

            if account_col_idx is None or name_col_idx is None:
                raise ValueError(
                    f"Required columns not found. Need account number and name columns. "
                    f"Found headers: {headers}"
                )

            balance_col_idx = self._find_column_index(normalized_headers, self.BALANCE_COLUMNS)
            debit_col_idx = self._find_column_index(normalized_headers, self.DEBIT_COLUMNS)
            credit_col_idx = self._find_column_index(normalized_headers, self.CREDIT_COLUMNS)

            if balance_col_idx is None and debit_col_idx is None and credit_col_idx is None:
                raise ValueError(
                    "No balance columns found. Need either 'balance' or 'debit'/'credit' columns"
                )

            # Parse data rows
            accounts = []
            header_row_num = header_row[0].row
            for row in sheet.iter_rows(min_row=header_row_num + 1):
                account = self._parse_excel_row(
                    row, account_col_idx, name_col_idx, balance_col_idx, debit_col_idx, credit_col_idx
                )
                if account:
                    accounts.append(account)

            workbook.close()

            return SaldenlisteData(
                tax_year=tax_year,
                accounts=accounts,
                confidence=self._calculate_confidence(accounts),
                format_detected=format_detected,
            )

        except Exception as e:
            raise ValueError(f"Failed to parse Excel file: {str(e)}") from e

    def detect_format(self, file_path: str) -> SaldenlisteFormat:
        """
        Auto-detect Saldenliste format (BMD, RZL, custom).

        Args:
            file_path: Path to the file

        Returns:
            Detected format

        Raises:
            ValueError: If file cannot be read
        """
        path = Path(file_path)
        if not path.exists():
            raise ValueError(f"File not found: {file_path}")

        suffix = path.suffix.lower()

        try:
            if suffix == ".csv":
                reader = self._open_csv_reader(file_path)
                if reader.fieldnames:
                    headers = [col.lower().strip() for col in reader.fieldnames]
                    return self._detect_format_from_headers(headers)

            elif suffix in [".xlsx", ".xls"]:
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = workbook.active
                header_row = self._find_header_row(sheet)
                if header_row:
                    headers = [
                        str(cell.value).lower().strip() if cell.value else ""
                        for cell in header_row
                    ]
                    workbook.close()
                    return self._detect_format_from_headers(headers)
                workbook.close()

        except Exception:
            pass

        return SaldenlisteFormat.UNKNOWN

    def normalize_account_number(self, account_number: str) -> str:
        """
        Normalize account number to standard format.

        Args:
            account_number: Raw account number from file

        Returns:
            Normalized account number (4 digits, zero-padded)
        """
        if not account_number:
            return ""

        # Remove non-numeric characters
        cleaned = re.sub(r"[^\d]", "", str(account_number))

        if not cleaned:
            return account_number  # Return original if no digits found

        # Pad to 4 digits if shorter
        if len(cleaned) < 4:
            cleaned = cleaned.zfill(4)

        return cleaned

    def extract_kontenklasse(self, account_number: str) -> Optional[int]:
        """
        Extract Kontenklasse (first digit) from account number.

        Args:
            account_number: Normalized account number

        Returns:
            Kontenklasse (0-9) or None if invalid
        """
        normalized = self.normalize_account_number(account_number)
        if normalized and normalized[0].isdigit():
            return int(normalized[0])
        return None

    # Private helper methods

    def _find_column(self, normalized_headers: Dict[str, str], candidates: List[str]) -> Optional[str]:
        """Find matching column name from candidates."""
        for candidate in candidates:
            if candidate in normalized_headers:
                return normalized_headers[candidate]
        return None

    def _find_column_index(self, normalized_headers: Dict[str, int], candidates: List[str]) -> Optional[int]:
        """Find matching column index from candidates."""
        for candidate in candidates:
            if candidate in normalized_headers:
                return normalized_headers[candidate]
        return None

    def _detect_format_from_headers(self, headers: List[str]) -> SaldenlisteFormat:
        """Detect format based on header column names."""
        headers_str = " ".join(headers).lower()

        # BMD typically has specific column patterns
        if "bmd" in headers_str or ("kto" in headers_str and "bezeichnung" in headers_str):
            return SaldenlisteFormat.BMD

        # RZL typically has different patterns
        if "rzl" in headers_str or ("kontonr" in headers_str and "kontobezeichnung" in headers_str):
            return SaldenlisteFormat.RZL

        # If we can identify standard columns, it's a custom format
        has_account = any(col in headers_str for col in ["konto", "account"])
        has_balance = any(col in headers_str for col in ["saldo", "balance", "betrag"])

        if has_account and has_balance:
            return SaldenlisteFormat.CUSTOM

        return SaldenlisteFormat.UNKNOWN

    def _find_header_row(self, sheet: Worksheet) -> Optional[List]:
        """Find the header row in an Excel sheet (may not be first row)."""
        # Check first 10 rows for headers
        for row in sheet.iter_rows(max_row=10):
            # Header row typically has text values in multiple columns
            non_empty = [cell for cell in row if cell.value is not None]
            if len(non_empty) >= 2:
                # Check if values look like headers (strings, not numbers)
                if all(isinstance(cell.value, str) for cell in non_empty[:3]):
                    return row
        return None

    def _open_csv_reader(self, file_path: str) -> csv.DictReader:
        """Open CSV files with a small encoding fallback set for Austrian exports."""
        raw_bytes = Path(file_path).read_bytes()
        for encoding in self.CSV_ENCODINGS:
            try:
                return csv.DictReader(io.StringIO(raw_bytes.decode(encoding)))
            except UnicodeDecodeError:
                continue
        raise ValueError("Unable to decode CSV file with supported encodings")

    def _parse_csv_row(
        self,
        row: Dict[str, str],
        account_col: str,
        name_col: str,
        balance_col: Optional[str],
        debit_col: Optional[str],
        credit_col: Optional[str],
    ) -> Optional[AccountBalance]:
        """Parse a single CSV row into AccountBalance."""
        account_number = row.get(account_col, "").strip()
        account_name = row.get(name_col, "").strip()

        if not account_number or not account_name:
            return None

        # Normalize account number
        normalized_account = self.normalize_account_number(account_number)
        kontenklasse = self.extract_kontenklasse(normalized_account)

        # Parse balance values
        debit_balance = None
        credit_balance = None
        balance = Decimal("0")

        try:
            if balance_col:
                balance_str = row.get(balance_col, "0").strip()
                balance = self._parse_decimal(balance_str)
            else:
                if debit_col:
                    debit_str = row.get(debit_col, "0").strip()
                    debit_balance = self._parse_decimal(debit_str)
                if credit_col:
                    credit_str = row.get(credit_col, "0").strip()
                    credit_balance = self._parse_decimal(credit_str)

                # Calculate net balance (debit - credit)
                balance = (debit_balance or Decimal("0")) - (credit_balance or Decimal("0"))

        except (ValueError, InvalidOperation):
            # Skip rows with invalid balance values
            return None

        return AccountBalance(
            account_number=normalized_account,
            account_name=account_name,
            debit_balance=debit_balance,
            credit_balance=credit_balance,
            balance=balance,
            kontenklasse=kontenklasse,
        )

    def _parse_excel_row(
        self,
        row: tuple,
        account_col_idx: int,
        name_col_idx: int,
        balance_col_idx: Optional[int],
        debit_col_idx: Optional[int],
        credit_col_idx: Optional[int],
    ) -> Optional[AccountBalance]:
        """Parse a single Excel row into AccountBalance."""
        account_number = str(row[account_col_idx].value or "").strip()
        account_name = str(row[name_col_idx].value or "").strip()

        if not account_number or not account_name:
            return None

        # Normalize account number
        normalized_account = self.normalize_account_number(account_number)
        kontenklasse = self.extract_kontenklasse(normalized_account)

        # Parse balance values
        debit_balance = None
        credit_balance = None
        balance = Decimal("0")

        try:
            if balance_col_idx is not None:
                balance_value = row[balance_col_idx].value
                balance = self._parse_decimal(balance_value)
            else:
                if debit_col_idx is not None:
                    debit_value = row[debit_col_idx].value
                    debit_balance = self._parse_decimal(debit_value)
                if credit_col_idx is not None:
                    credit_value = row[credit_col_idx].value
                    credit_balance = self._parse_decimal(credit_value)

                # Calculate net balance (debit - credit)
                balance = (debit_balance or Decimal("0")) - (credit_balance or Decimal("0"))

        except (ValueError, InvalidOperation):
            # Skip rows with invalid balance values
            return None

        return AccountBalance(
            account_number=normalized_account,
            account_name=account_name,
            debit_balance=debit_balance,
            credit_balance=credit_balance,
            balance=balance,
            kontenklasse=kontenklasse,
        )

    def _parse_decimal(self, value: Any) -> Decimal:
        """Parse a value into Decimal, handling various formats."""
        if value is None or value == "":
            return Decimal("0")

        if isinstance(value, (int, float)):
            return Decimal(str(value))

        if isinstance(value, str):
            # Remove common formatting (spaces, currency symbols)
            cleaned = value.strip().replace(" ", "").replace("€", "").replace("EUR", "")

            # Handle German number format (comma as decimal separator)
            if "," in cleaned and "." in cleaned:
                if cleaned.rfind(",") > cleaned.rfind("."):
                    # Format like 1.234,56 (German)
                    cleaned = cleaned.replace(".", "").replace(",", ".")
                else:
                    # Format like 1,234.56 (US)
                    cleaned = cleaned.replace(",", "")
            elif "," in cleaned:
                # Format like 1234,56 (German)
                cleaned = cleaned.replace(",", ".")

            # Remove any remaining non-numeric characters except decimal point and minus
            cleaned = re.sub(r"[^\d.-]", "", cleaned)

            if not cleaned or cleaned == "-":
                return Decimal("0")

            return Decimal(cleaned)

        return Decimal("0")

    def _calculate_confidence(self, accounts: List[AccountBalance]) -> float:
        """
        Calculate confidence score for parsed data.

        Args:
            accounts: List of parsed accounts

        Returns:
            Confidence score (0.0 to 1.0)
        """
        if not accounts:
            return 0.0

        score = 0.0
        total_checks = 0

        # Check 1: All accounts have valid account numbers
        valid_accounts = sum(1 for acc in accounts if acc.account_number and len(acc.account_number) >= 4)
        score += (valid_accounts / len(accounts)) * 0.3
        total_checks += 0.3

        # Check 2: All accounts have Kontenklasse assigned
        with_kontenklasse = sum(1 for acc in accounts if acc.kontenklasse is not None)
        score += (with_kontenklasse / len(accounts)) * 0.3
        total_checks += 0.3

        # Check 3: All accounts have non-empty names
        with_names = sum(1 for acc in accounts if acc.account_name)
        score += (with_names / len(accounts)) * 0.2
        total_checks += 0.2

        # Check 4: At least some accounts have non-zero balances
        with_balance = sum(1 for acc in accounts if acc.balance != Decimal("0"))
        if with_balance > 0:
            score += 0.2
        total_checks += 0.2

        return min(score / total_checks, 1.0) if total_checks > 0 else 0.0
